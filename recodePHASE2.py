#Lauren Magee
#Program 3

#Recode Part 2
#Took the XLSX, modified the column values to One Hot or Standardized, then appended all new recoded values to a final XLSX
#This code was editted to produce 3 XLSX files, one with 70% of data for training, one with 15% for testing, and the last 15% for
#validation

from openpyxl import load_workbook
import xlsxwriter
from scipy import stats
from sklearn.preprocessing import LabelEncoder

workbook = load_workbook(filename="D:\Documents\CS461/output.xlsx")
sheet = workbook.active

#Gender Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999,min_col=2,max_col=2, values_only=True):
    value = list(value)
    tempArray.append(value)

genderArray = []

for value in tempArray:
    for value1 in value:
        if value1 == "Male":
            value1 = 0
            genderArray.append(value1)
        elif value1 == "Female":
            value1 = 1
            genderArray.append(value1)

#Driver's License Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999,min_col=4,max_col=4, values_only=True):
    value = list(value)
    tempArray.append(value)

driversLicense = []

for value in tempArray:
    for value1 in value:
        driversLicense.append(value1)

#Previously Insured Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=6,max_col=6, values_only=True):
    value = list(value)
    tempArray.append(value)

prevInsured = []

for value in tempArray:
    for value1 in value:
        prevInsured.append(value1)

#Vehicle Damage Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=8,max_col=8, values_only=True):
    value = list(value)
    tempArray.append(value)

carDamage = []

for value in tempArray:
    for value1 in value:
        if value1 == "Yes":
            value1 = 1
            carDamage.append(value1)
        elif value1 == "No":
            value1 = 0
            carDamage.append(value1)

#Age Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999,min_col=3,max_col=3, values_only=True):
    value = list(value)
    tempArray.append(value)

ageArray = []

for value in tempArray:
    for argument in value:
        if argument == '16' or argument == '17' or argument == '18' or argument == '19' or argument == '20':
            argument = 0
        elif argument == '21' or argument == '22' or argument == '23' or argument == '24' or argument == '25':
            argument = 1
        elif argument == '26' or argument == '27' or argument == '28' or argument == '29' or argument == '30':
            argument = 2
        elif argument == '31' or argument == '32' or argument == '33' or argument == '34' or argument == '35':
            argument = 3
        elif argument == '36' or argument == '37' or argument == '38' or argument == '39' or argument == '40':
            argument = 4
        elif argument == '41' or argument == '42' or argument == '43' or argument == '44' or argument == '45':
            argument = 5
        elif argument == '46' or argument == '47' or argument == '48' or argument == '49' or argument == '50':
            argument = 6
        elif argument == '51' or argument == '52' or argument == '53' or argument == '54' or argument == '55':
            argument = 7
        elif argument == '56' or argument == '57' or argument == '58' or argument == '59' or argument == '60':
            argument = 8
        elif argument == '61' or argument == '62' or argument == '63' or argument == '64' or argument == '65':
            argument = 9
        elif argument == '66' or argument == '67' or argument == '68' or argument == '69' or argument == '70':
            argument = 10
        elif argument == '71' or argument == '72' or argument == '73' or argument == '74' or argument == '75':
            argument = 11
        elif argument == '76' or argument == '77' or argument == '78' or argument == '79' or argument == '80':
            argument = 12
        elif argument == '81' or argument == '82' or argument == '83' or argument == '84' or argument == '85':
            argument = 13
        ageArray.append(argument)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(ageArray)
ageArray = label_ids

#Region Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=5,max_col=5, values_only=True):
    value = list(value)
    tempArray.append(value)

region = []

for value in tempArray:
    for value1 in value:
        region.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(region)
region = label_ids

#Yearly Premium Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=9,max_col=9, values_only=True):
    value = list(value)
    tempArray.append(value)

yearPremium = []

for value in tempArray:
    for value1 in value:
        value1 = int(value1)
        yearPremium.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(yearPremium)
yearPremium = stats.zscore(yearPremium)

#Policy Sales Channel Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=10,max_col=10, values_only=True):
    value = list(value)
    tempArray.append(value)

policySales = []

for value in tempArray:
    for value1 in value:
        policySales.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(policySales)
policySales = label_ids

#Vintage Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999,min_col=11,max_col=11, values_only=True):
    value = list(value)
    tempArray.append(value)

vintage = []

for value in tempArray:
    for value1 in value:
        value1 = int(value1)
        vintage.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(vintage)
vintage = stats.zscore(vintage)


#Vehicle Age Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=7,max_col=7, values_only=True):
    value = list(value)
    tempArray.append(value)

carAge = []

for value in tempArray:
    for value1 in value:
        if value1 == "> 2 Years":
            value1 = 0
        if value1 == "1-2 Year":
            value1=1
        if value1 == "< 1 Year":
            value1 = 2
        carAge.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(carAge)
carAge = label_ids

#Response Recode
tempArray = []
for value in sheet.iter_cols(min_row=2, max_row=265999, min_col=12,max_col=12, values_only=True):
    value = list(value)
    tempArray.append(value)

response = []

for value in tempArray:
    for value1 in value:
        response.append(value1)

label_encoder=LabelEncoder()
label_ids=label_encoder.fit_transform(response)
response = label_ids

###############################################################################################################################
workbook = xlsxwriter.Workbook('D:\Documents\CS461/FINALtraining70.xlsx')
worksheet = workbook.add_worksheet()

# Start from the first cell.
# Rows and columns are zero indexed.
row = 0
column = 0

# iterating through content list
for item in genderArray:
    worksheet.write(row, 0, item)
    row += 1

row = 0
for item in driversLicense:
    worksheet.write(row, 2, item)
    row += 1

row = 0
for item in prevInsured:
    worksheet.write(row, 4, item)
    row += 1

row = 0
for item in carAge:
    worksheet.write(row, 5, item)
    row += 1

row = 0
for item in carDamage:
    worksheet.write(row, 6, item)
    row += 1

row = 0
for item in yearPremium:
    worksheet.write(row, 7, item)
    row += 1

row = 0
for item in policySales:
    worksheet.write(row, 8, item)
    row += 1

row = 0
for item in vintage:
    worksheet.write(row, 9, item)
    row += 1

row = 0
for item in response:
    worksheet.write(row, 10, item)
    row += 1

row = 0
for item in ageArray:
    worksheet.write(row, 1, item)
    row += 1

row = 0
for item in region:
    worksheet.write(row, 3, item)
    row += 1

workbook.close()